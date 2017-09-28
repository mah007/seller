import openpyxl
from random import randint
from config import ExcelImporter

class ImportExcel(object):

    def __init__(self):
        self.file_path = ExcelImporter.FILE_PATH

    def open_work_sheet(self, sheet_name):
        work_book = openpyxl.load_workbook(self.file_path, data_only=True)
        work_sheet = work_book.get_sheet_by_name(sheet_name)
        return work_sheet

    def get_column_by_letter(self, letter):
        return openpyxl.utils.cell.column_index_from_string(letter) - 1

    def getGeneralExchangeFromWorkbook(self, sheet_name):
        datas = []
        try:
            work_sheet = self.open_work_sheet(sheet_name)
            for index, current_row in enumerate(work_sheet.rows):
                if index > 3:
                    order_number = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_ORDER_NUMBER)].value
                    sku = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_SKU)].value
                    item_status = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_ITEM_STATUS)].value
                    tracking_number = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_TRACKING_NUMBER)].value
                    shipment_type = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_SHIPMENT_TYPE)].value
                    payment_method = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_PAYMENT_METHOD)].value
                    sales_deliver = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_SALES_DELIVER)].value
                    sales_return = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_SALES_RETURN)].value
                    wrong_status = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_WRONG_STATUS)].value
                    seller_delivery = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_SELLER_DELIVERY)].value
                    product_bundling = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_PRODUCT_BUNDING)].value
                    subsidy = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_SUBSIDY)].value
                    commission = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_COMISSION)].value
                    sum_of_fee = current_row[self.get_column_by_letter(ExcelImporter.COLUMN_SUM_OF_FEE)].value

                    if(order_number != None):
                        datas.append({
                            'order_number': order_number,
                            'sku': sku,
                            'item_status': item_status,
                            'tracking_number': tracking_number,
                            'shipment_type': shipment_type,
                            'payment_method': payment_method,
                            'sales_deliver': sales_deliver,
                            'sales_return': sales_return,
                            'wrong_status': wrong_status,
                            'seller_delivery': seller_delivery,
                            'product_bundling': product_bundling,
                            'subsidy': subsidy,
                            'commission': commission,
                            'sum_of_fee': sum_of_fee
                        })

            return datas
        except Exception as ex:
            raise ex
            return datas

    def getGeneralExchange(self):
        general_exchange = self.getGeneralExchangeFromWorkbook(ExcelImporter.GENERAL_EXCHANGE)
        return general_exchange

